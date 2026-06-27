#!/usr/bin/env python3
"""
DRY-RUN PREVIEW — Historical challan import
Parses Soybean Dispatch 2026.xls and shows every row that would be inserted.
NO DB WRITES — read-only preview.
"""

import xlrd
import openpyxl
from datetime import datetime, date
from collections import defaultdict, Counter

EXCEL_PATH  = "/Users/radhikavarma/Downloads/Soybean Dispatch - 2026.xls"
DIST_XLSX   = "/Users/radhikavarma/Desktop/ASN-Challan/distributors_clean.xlsx"
RETAIL_XLSX = "/Users/radhikavarma/Desktop/ASN-Challan/retailers_clean.xlsx"

ASN_MAX_DC   = 231
ASIAN_MAX_DC = 320

# Known data fixes: (company, dc_number, product_name) → corrected qty_qtl
# ASN DC 107 Asian-777: 9.90 qtl is a typo — should be 9.99 (37 bags × 27kg)
DATA_FIXES = {
    ("ASN", 107, "Asian-777"): 9.99,
}

# ── Product packing sizes (kg per bag) ────────────────────────────────────────
# Verified via arithmetic (qty_qtl * 100 / packing_kg must give integer bags)
# and cross-checked against user-confirmed values from the DB.
PACKING = {
    "JS-335":           30.0,   # user confirmed 30kg
    "JS-9305":          30.0,   # verified via arithmetic
    "JS-9560":          30.0,   # verified via arithmetic
    "RVSM-1135":        30.0,   # 1 data point; likely 30kg
    "KDS-726":          30.0,   # no data in import range; likely 30kg
    "Asian-777":        27.0,   # verified via arithmetic (DC107 has 1 typo: 9.90 should be 9.99)
    "Asian Sanjivani":  27.0,   # verified via arithmetic
    "Confidence":       27.0,   # verified via arithmetic (all 93 rows)
    "Krushi Gold-151":  25.0,   # user confirmed 25kg
    "Game Changer":     25.0,   # verified via arithmetic (all 84 rows)
}

# Column layout (same for both sheets)
COL_SNO    = 0
COL_DATE   = 1
COL_DC     = 2
COL_PARTY  = 3
COL_PLACE  = 4
COL_QTY    = 5
COL_TRUCK  = -3   # 3rd from end
COL_TRANS  = -2   # 2nd from end
COL_BILTY  = -1   # last

VARIETY_COLS_ASN   = {6:"JS-335", 7:"JS-9305", 8:"JS-9560",
                      9:"RVSM-1135", 10:"KDS-726", 11:"Asian-777",
                      12:"Asian Sanjivani", 13:"Confidence"}
VARIETY_COLS_ASIAN = {6:"JS-335", 7:"JS-9305", 8:"JS-9560",
                      9:"Krushi Gold-151", 10:"Game Changer"}

# ── Load distributor list ──────────────────────────────────────────────────────
wb_dist = openpyxl.load_workbook(DIST_XLSX)
ws_dist = wb_dist.active
dist_list = []
for row in ws_dist.iter_rows(min_row=2, values_only=True):
    if row[0]:
        dist_list.append({"name": row[0].strip(), "city": row[1] or ""})
dist_upper_to_name = {d["name"].upper(): d["name"] for d in dist_list}

# Load retailer list (existing)
wb_ret = openpyxl.load_workbook(RETAIL_XLSX)
ws_ret = wb_ret.active
ret_cols = [ws_ret.cell(1, c).value for c in range(1, ws_ret.max_column + 1)]
existing_retailers = set()  # (retailer_name_upper, distributor_name_upper)
# retailers_clean.xlsx columns: name, city, phone, distributor_name, address
for row in ws_ret.iter_rows(min_row=2, values_only=True):
    if row[0]:
        ret_name = str(row[0]).strip().upper()
        dist_name = str(row[3]).strip().upper() if len(row) > 3 and row[3] else ""
        existing_retailers.add((ret_name, dist_name))

# ── Helper functions ───────────────────────────────────────────────────────────
def cell_str(cell):
    if cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        v = cell.value
        if v == int(v):
            return str(int(v))
        return str(v)
    return str(cell.value).strip()

def cell_num(cell):
    if cell.ctype == xlrd.XL_CELL_NUMBER:
        return cell.value
    try:
        return float(str(cell.value).strip())
    except:
        return 0.0

def parse_date(cell, wb):
    """Parse date from xlrd cell. Handles float serial and DD.MM.YY string."""
    if cell.ctype == xlrd.XL_CELL_DATE or (cell.ctype == xlrd.XL_CELL_NUMBER and cell.value > 1000):
        try:
            t = xlrd.xldate_as_tuple(cell.value, wb.datemode)
            return date(t[0], t[1], t[2])
        except:
            pass
    s = str(cell.value).strip()
    for fmt in ["%d.%m.%y", "%d.%m.%Y", "%d/%m/%y", "%d/%m/%Y", "%d-%m-%Y", "%d-%b-%Y"]:
        try:
            return datetime.strptime(s, fmt).date()
        except:
            pass
    return None

# Manual mapping for Excel name → DB name (spelling variations in Excel vs distributors_clean.xlsx)
EXCEL_DIST_MAP = {
    "SAI KRUPA KRISHI VIKAS KENDRA":   "SAI KRUPA KVK",
    "JAI KISAN KRISHI SEVA KENDRA":    "Jai Kisan Krishi Kendra",
    "JAISWAL KRISHI SEVA KENDRA":      "Jaiswal Krishi Kendra",
    "RAMLINGESHWAR FERTILIZER":        "Ramlingeshwar Fertilizers",
    "GIRIRAJ TRADING CO.":             "GIRIRAJ TRADING CO.",
    "SHRIPAD KRISHI KENDRA":           "Sripad Krishi Seva Kendra",
    "KOTKAR AGRO TRADERS":             "Kotkar Agro",
    "RAM KRISHI KENDRA":               "Ram Krushi Kendra",
    "GANESH KRISHI VIKAS KENDRA":      "Ganesh Krushi Vikas Kendra",
    "ASHOK KUMAR BAJRANGLAL":          "Ashok Kumar Bajranglal",
    "VIVAN FERTILIZERS":               "Vivaan Fertilizer",
    "VIVAN FERTILIZER":                "Vivaan Fertilizer",
    "TIRUPATI AGRO AGENCIES":          "Tirupati Agro Agency",
    "PADMA AGRO TRADERS":              "Padma Agro",
    "NALLAWAR KRISHI KENDRA":          "Nallawar Krishi Kendra Patan",
    "VARUN SEEDS & FERTILIZERS":       "Varun Seeds And Fertilizers",
    "VARUN SEEDS & FERTILIZER":        "Varun Seeds And Fertilizers",
    "N.C. KOCHAR":                     "N C Kochar",
    "SHIVSHANKAR KIRANA & FERTILIZER": "Shivshankar Kirana & Ferti.",
    "PRANAV AGENCY":                   "Pranav Agencies",
    "KHANDESHWAR KRUSHI KENDRA":       "Khandeshwar Krushi Kendra",
    "KARANJA SAHKARI TALUKA":          "Karanja Sahkari Taluka Kharedi Vikri",
    # truly missing from DB — will appear as UNKNOWN_DIST in output
    "CHARBHUJA ENTERPRISES":           "CHARBHUJA ENTERPRISES [NOT IN DB]",
}

def match_distributor(header_col0):
    """Extract distributor name from section header like 'BALAJI KRISHI SEVA KENDRA- NANDED'."""
    raw = str(header_col0).strip()
    # Strip city suffix: "NAME- CITY" or "NAME -CITY"
    name_part = raw
    for sep in ["- ", " -", "-"]:
        if sep in raw:
            name_part = raw.split(sep)[0].strip()
            break

    name_up = name_part.strip().upper()

    # 1. Check manual map first
    if name_up in EXCEL_DIST_MAP:
        mapped = EXCEL_DIST_MAP[name_up]
        if "[NOT IN DB]" in mapped:
            return mapped  # pass through with warning marker
        return mapped

    # 2. Exact match in DB
    if name_up in dist_upper_to_name:
        return dist_upper_to_name[name_up]

    # 3. Try full raw string
    raw_up = raw.upper()
    if raw_up in dist_upper_to_name:
        return dist_upper_to_name[raw_up]

    # 4. Partial match (safe: only if one side fully contains the other)
    for d_up, d_name in dist_upper_to_name.items():
        if d_up in raw_up or raw_up in d_up:
            return d_name

    return None

# Known aliases: party names that ARE the distributor (stored under a different name in DB)
DIST_ALIASES = {
    "KHAREDI VIKRI": {"KARANJA SAHKARI TALUKA"},
}

def _norm(s):
    """Normalize common Hindi transliteration variants (KRUSHI/KRISHI only)."""
    return s.replace("KRUSHI", "KRISHI")

def is_same_as_dist(party_name, dist_name):
    """True if the party name refers to the current distributor (ship-to-dist)."""
    if not party_name:
        return True
    pu_raw = party_name.strip().upper()
    du_raw = dist_name.strip().upper()
    pu = _norm(pu_raw)
    du = _norm(du_raw)
    if pu == du:
        return True
    # All words in party are subset of dist words (handles subset names)
    p_words = set(pu.split())
    d_words = set(du.split())
    if p_words <= d_words:
        return True
    # KSK abbreviation: party starts with same first word as dist
    if pu_raw.endswith(" KSK") or pu_raw.endswith(" KK"):
        p_first = pu_raw.split()[0]
        d_first = du_raw.split()[0]
        if p_first == d_first:
            return True
    # Abbreviated or spelling-variant form: same first significant word + party ≤ dist length
    p_sig = [w for w in pu.split() if len(w) > 2]
    d_sig = [w for w in du.split() if len(w) > 2]
    if p_sig and d_sig and p_sig[0] == d_sig[0]:
        if len(pu) < len(du):
            return True
        if len(pu) == len(du):
            # Same length — require 2nd significant word to also match (spelling variant)
            if len(p_sig) < 2 or len(d_sig) < 2 or p_sig[1] == d_sig[1]:
                return True
    return False

# ── Parse sheets ──────────────────────────────────────────────────────────────
wb = xlrd.open_workbook(EXCEL_PATH)

SHEETS = [
    ("ASN Soya Final",   "ASN",   ASN_MAX_DC,   VARIETY_COLS_ASN),
    ("ASIAN Soya Final", "ASIAN", ASIAN_MAX_DC, VARIETY_COLS_ASIAN),
]

# Results
all_challans  = []   # {company, dc_number, dc_date, dist_name, party_name, ship_to_dist,
                     #  total_qty_qtl, lorry_no, transport, lr_no, items[]}
new_retailers = []   # {name, city, dist_name}
warnings      = []
transfers     = []   # {company, dc_num, party_name, dist_name}

for sheet_name, comp_code, max_dc, variety_cols in SHEETS:
    ws = wb.sheet_by_name(sheet_name)
    ncols = ws.ncols

    current_dist = None

    for ri in range(ws.nrows):
        row = ws.row(ri)
        c0 = cell_str(row[0])
        c0_up = c0.strip().upper()

        import re as _re

        # ── Section header detection ──────────────────────────────────────────
        # Rule: col0 has text (non-numeric), col1-col5 all empty/zero
        # Exception: "62A"-style S.No. means this is a suffixed data row.
        if row[0].ctype not in (xlrd.XL_CELL_NUMBER, xlrd.XL_CELL_EMPTY):
            # Check if col0 looks like a suffixed serial number (e.g. "62A", "1A")
            _sno_m = _re.match(r'^(\d+)[A-Za-z]$', c0.strip())
            if not _sno_m:
                # True section header: col1-col5 all empty/zero
                others_empty = all(
                    row[ci].ctype == xlrd.XL_CELL_EMPTY or str(row[ci].value).strip() in ("", "0", "0.0")
                    for ci in range(1, min(6, ncols))
                )
                if others_empty and c0:
                    if c0_up in ("ASN AGRI GENETICS PVT.LTD.", "ASIAN SEEDS PVT.LTD.",
                                 "ASN AGRI GENETICS PVT. LTD.", "ASIAN SEEDS PVT. LTD."):
                        continue
                    if c0_up.startswith("DISPATCH DETAILS"):
                        continue
                    dist_name = match_distributor(c0)
                    if dist_name:
                        current_dist = dist_name
                    else:
                        warnings.append(f"[{sheet_name} R{ri+1}] Unmatched section: '{c0}'")
                continue
            # Fall through to data row handling with the suffixed sno

        # ── Data row detection ────────────────────────────────────────────────
        # col0 is S.No. (numeric or "62A"-style text), col2 is DC No. (integer or "107A")
        try:
            sno = int(float(c0)) if row[0].ctype == xlrd.XL_CELL_NUMBER else int(_re.match(r'^(\d+)', c0).group(1))
        except:
            continue
        dc_raw = cell_str(row[COL_DC]).strip()
        _m = _re.match(r'^(\d+)\s*([A-Za-z]?)$', dc_raw)
        if not _m:
            continue
        dc_num    = int(_m.group(1))
        dc_suffix = _m.group(2) or None

        # Skip if DC out of import range
        if dc_num > max_dc:
            continue

        if not current_dist:
            warnings.append(f"[{sheet_name} DC {dc_num}] No distributor context at R{ri+1}")
            continue

        # Parse date
        dc_date = parse_date(row[COL_DATE], wb)
        if dc_date is None:
            warnings.append(f"[{sheet_name} DC {dc_num}] Bad date '{cell_str(row[COL_DATE])}', defaulting to 2026-04-01")
            dc_date = date(2026, 4, 1)

        party_name  = cell_str(row[COL_PARTY]).strip()
        place_name  = cell_str(row[COL_PLACE]).strip()
        total_qty   = cell_num(row[COL_QTY])

        # Truck, transport, bilty (last 3 cols before freight cols)
        lorry_no  = cell_str(row[ncols + COL_TRUCK]).strip()  if ncols >= 3 else ""
        transport = cell_str(row[ncols + COL_TRANS]).strip()  if ncols >= 2 else ""
        lr_no     = cell_str(row[ncols + COL_BILTY]).strip()  if ncols >= 1 else ""

        # Detect transfer rows (ADD/LESS adjustments for OM AGRO)
        party_up = party_name.upper()
        if "ADD:-" in party_up or "LESS:-" in party_up:
            transfers.append({"company": comp_code, "dc_num": dc_num,
                               "party": party_name, "dist": current_dist})
            continue

        # Determine ship-to
        ship_to_dist = is_same_as_dist(party_name, current_dist)

        # Determine retailer / new retailer flag
        new_ret = False
        if not ship_to_dist and party_name:
            ret_key = (party_up, current_dist.upper())
            if ret_key not in existing_retailers:
                new_ret = True
                if not any(nr["name"].upper() == party_up and nr["dist_name"].upper() == current_dist.upper()
                           for nr in new_retailers):
                    new_retailers.append({
                        "name": party_name,
                        "city": place_name,
                        "dist_name": current_dist
                    })

        # Build line items
        items = []
        for ci, variety_token in variety_cols.items():
            if ci >= ncols:
                continue
            qty_val = cell_num(row[ci])
            if qty_val <= 0:
                continue
            # Apply known data fixes
            fix_key = (comp_code, dc_num, variety_token)
            if fix_key in DATA_FIXES:
                qty_val = DATA_FIXES[fix_key]
            pkg = PACKING.get(variety_token, 30.0)
            bags_calc = round((qty_val * 100) / pkg)  # calculated for display
            items.append({
                "product_name": variety_token,
                "qty_qtl":      round(qty_val, 2),
                "bags":         bags_calc,
                "packing_kg":   pkg,
            })

        all_challans.append({
            "company":       comp_code,
            "dc_number":     dc_num,
            "dc_suffix":     dc_suffix,
            "dc_date":       dc_date.isoformat(),
            "dist_name":     current_dist,
            "party_name":    party_name,
            "place":         place_name,
            "ship_to_dist":  ship_to_dist,
            "new_retailer":  new_ret,
            "total_qty_qtl": round(total_qty, 2),
            "lorry_no":      lorry_no,
            "transport":     transport,
            "lr_no":         lr_no,
            "items":         items,
        })

# ── Sort challans by company then DC number ──────────────────────────────────
all_challans.sort(key=lambda c: (c["company"], c["dc_number"]))

# ── PRINT REPORT ──────────────────────────────────────────────────────────────
SEP = "─" * 120
DSEP = "═" * 120

print(DSEP)
print("  DRY-RUN IMPORT PREVIEW — Soybean Dispatch 2026")
print(DSEP)
asn_count   = sum(1 for c in all_challans if c["company"] == "ASN")
asian_count = sum(1 for c in all_challans if c["company"] == "ASIAN")
ship_to_d   = sum(1 for c in all_challans if c["ship_to_dist"])
ship_to_r   = sum(1 for c in all_challans if not c["ship_to_dist"])
item_count  = sum(len(c["items"]) for c in all_challans)
print(f"  Challans to insert:     {len(all_challans)}  (ASN={asn_count}, ASIAN={asian_count})")
print(f"  Ship-to-distributor:    {ship_to_d}")
print(f"  Ship-to-retailer:       {ship_to_r}")
print(f"  Challan items to insert:{item_count}")
print(f"  New retailers to create:{len(new_retailers)}")
print(f"  Transfer rows (skipped):{len(transfers)}")
print(f"  Warnings:               {len(warnings)}")
print()

# ── NEW RETAILERS ─────────────────────────────────────────────────────────────
if new_retailers:
    print(SEP)
    print(f"  NEW RETAILERS  (would be created in retailers table before import)")
    print(SEP)
    print(f"  {'#':<4}  {'Retailer Name':<40}  {'City':<18}  {'Under Distributor'}")
    print(f"  {'─'*4}  {'─'*40}  {'─'*18}  {'─'*35}")
    for i, nr in enumerate(new_retailers, 1):
        print(f"  {i:<4}  {nr['name'][:39]:<40}  {nr['city'][:17]:<18}  {nr['dist_name']}")
    print()

# ── CHALLANS TABLE ────────────────────────────────────────────────────────────
print(SEP)
print("  CHALLANS TABLE")
print(SEP)
print(f"  {'Co':<6}  {'DC#':<5}  {'Date':<11}  {'Distributor':<38}  {'Ship-To (Party)':<38}  {'QtyQtl':>8}  {'Items':>5}  {'Lorry':<14}  {'Bilty'}")
print(f"  {'─'*6}  {'─'*5}  {'─'*11}  {'─'*38}  {'─'*38}  {'─'*8}  {'─'*5}  {'─'*14}  {'─'*8}")
for ch in all_challans:
    ship_label = "(→ DIST)" if ch["ship_to_dist"] else ch["party_name"]
    new_flag   = " [NEW]" if ch["new_retailer"] else ""
    dc_label = str(ch['dc_number']) + (ch['dc_suffix'] or '')
    print(f"  {ch['company']:<6}  {dc_label:<6}  {ch['dc_date']:<11}  "
          f"{ch['dist_name'][:37]:<38}  {(ship_label + new_flag)[:37]:<38}  "
          f"{ch['total_qty_qtl']:>8.2f}  {len(ch['items']):>5}  "
          f"{ch['lorry_no'][:13]:<14}  {ch['lr_no'][:8]}")

# ── CHALLAN ITEMS TABLE ───────────────────────────────────────────────────────
print()
print(SEP)
print("  CHALLAN_ITEMS TABLE  (handwrite mode: lot_id=NULL, lot_number='')")
print(SEP)
print(f"  {'Co':<6}  {'DC#':<5}  {'Product':<22}  {'qty_qtl':>8}  {'pkg_kg':>6}  {'bags':>5}  Note")
print(f"  {'─'*6}  {'─'*5}  {'─'*22}  {'─'*8}  {'─'*6}  {'─'*5}  {'─'*25}")
for ch in all_challans:
    for it in ch["items"]:
        kg_exact = round(it["qty_qtl"] * 100)
        note = "⚠ non-integer bags" if kg_exact % int(it["packing_kg"]) != 0 else ""
        print(f"  {ch['company']:<6}  {ch['dc_number']:<5}  {it['product_name']:<22}  "
              f"{it['qty_qtl']:>8.2f}  {it['packing_kg']:>6.0f}  {it['bags']:>5}  {note}")

# ── TRANSFER ROWS (SKIPPED) ───────────────────────────────────────────────────
if transfers:
    print()
    print(SEP)
    print("  TRANSFER ROWS — SKIPPED (import separately as D2D transfers)")
    print(SEP)
    for t in transfers:
        print(f"  [{t['company']}] DC {t['dc_num']}  | {t['dist']} | party='{t['party']}'")

# ── WARNINGS ─────────────────────────────────────────────────────────────────
if warnings:
    print()
    print(SEP)
    print("  WARNINGS")
    print(SEP)
    for w in warnings:
        print(f"  ! {w}")

# ── SUMMARY BY DISTRIBUTOR ────────────────────────────────────────────────────
print()
print(SEP)
print("  SUMMARY BY DISTRIBUTOR")
print(SEP)
by_dist = defaultdict(lambda: {"asn": 0, "asian": 0, "items": 0})
for ch in all_challans:
    key = ch["dist_name"]
    by_dist[key][ch["company"].lower()] += 1
    by_dist[key]["items"] += len(ch["items"])
print(f"  {'Distributor':<42}  {'ASN DCs':>8}  {'ASIAN DCs':>9}  {'Total DCs':>9}  {'Items':>6}")
print(f"  {'─'*42}  {'─'*8}  {'─'*9}  {'─'*9}  {'─'*6}")
for dist_name in sorted(by_dist.keys()):
    v = by_dist[dist_name]
    total = v["asn"] + v["asian"]
    print(f"  {dist_name:<42}  {v['asn']:>8}  {v['asian']:>9}  {total:>9}  {v['items']:>6}")
total_asn   = sum(v["asn"]   for v in by_dist.values())
total_asian = sum(v["asian"] for v in by_dist.values())
total_all   = total_asn + total_asian
total_items = sum(v["items"] for v in by_dist.values())
print(f"  {'TOTAL':<42}  {total_asn:>8}  {total_asian:>9}  {total_all:>9}  {total_items:>6}")

print()
print(DSEP)
print("  END OF DRY-RUN — NO DB WRITES PERFORMED")
print(DSEP)
